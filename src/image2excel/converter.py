"""Core conversion logic: image pixels to colored Excel cells."""

from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from PIL import Image

__all__ = ["im2xlsx"]

# Maximum dimensions when resizing (columns x rows)
MAX_WIDTH = 260
MAX_HEIGHT = 300


def _resize_image(
    image: Image.Image,
    *,
    keep_aspect: bool,
) -> Image.Image:
    """Resize an image to fit within MAX_WIDTH x MAX_HEIGHT."""
    width, height = image.size

    if width <= MAX_WIDTH and height <= MAX_HEIGHT:
        return image

    if keep_aspect:
        scale = min(MAX_WIDTH / width, MAX_HEIGHT / height)
        new_size = (int(width * scale), int(height * scale))
        image.thumbnail(new_size, Image.Resampling.LANCZOS)
        return image

    return image.resize((MAX_WIDTH, MAX_HEIGHT), Image.Resampling.LANCZOS)


def _pixel_to_hex(r: int, g: int, b: int) -> str:
    """Convert RGB values (0-255) to a hex color string without '#'."""
    return f"{r:02X}{g:02X}{b:02X}"


def _set_zoom(path: Path, zoom: int = 10) -> None:
    """Set the zoom level on all worksheets in an Excel file."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = zoom
    wb.save(path)


def im2xlsx(
    file: str | Path,
    *,
    resize: bool = True,
    keep_aspect: bool = False,
) -> Path:
    """Convert an image to an Excel file where each cell is a colored pixel.

    Args:
        file: Path to the source image.
        resize: Shrink images larger than 260x300 pixels. Defaults to True.
        keep_aspect: Preserve the aspect ratio when resizing. Defaults to False.

    Returns:
        Path to the generated ``.xlsx`` file.

    Raises:
        FileNotFoundError: If the image file does not exist.
        ValueError: If the image cannot be processed.
    """
    path = Path(file).resolve()
    if not path.is_file():
        msg = f"Image not found: {path}"
        raise FileNotFoundError(msg)

    image = Image.open(path).convert("RGB")

    if resize:
        image = _resize_image(image, keep_aspect=keep_aspect)

    pixels: np.ndarray = np.array(image)  # shape: (height, width, 3)
    height, width, _ = pixels.shape

    output_path = path.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    for row_idx in range(height):
        for col_idx in range(width):
            r, g, b = pixels[row_idx, col_idx]
            hex_color = _pixel_to_hex(int(r), int(g), int(b))
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Set narrow column widths so pixels look square-ish
    for col_idx in range(1, width + 1):
        letter = _col_letter(col_idx)
        ws.column_dimensions[letter].width = 2.0

    ws.sheet_view.zoomScale = 10
    wb.save(output_path)

    return output_path


def _col_letter(col: int) -> str:
    """Convert a 1-based column index to an Excel column letter (e.g. 1 -> A, 27 -> AA)."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result
