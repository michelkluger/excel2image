"""Tests for the image2excel converter."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook
from PIL import Image

from image2excel.converter import MAX_HEIGHT, MAX_WIDTH, _col_letter, _pixel_to_hex, im2xlsx


@pytest.fixture
def tiny_image(tmp_path: Path) -> Path:
    """Create a small 3x2 RGB image for testing."""
    pixels = np.array(
        [
            [[255, 0, 0], [0, 255, 0], [0, 0, 255]],
            [[255, 255, 0], [0, 255, 255], [255, 0, 255]],
        ],
        dtype=np.uint8,
    )
    img = Image.fromarray(pixels, "RGB")
    path = tmp_path / "test.png"
    img.save(path)
    return path


@pytest.fixture
def large_image(tmp_path: Path) -> Path:
    """Create an image larger than MAX_WIDTH x MAX_HEIGHT."""
    img = Image.new("RGB", (500, 600), color=(128, 64, 32))
    path = tmp_path / "big.png"
    img.save(path)
    return path


class TestPixelToHex:
    def test_red(self):
        assert _pixel_to_hex(255, 0, 0) == "FF0000"

    def test_green(self):
        assert _pixel_to_hex(0, 255, 0) == "00FF00"

    def test_black(self):
        assert _pixel_to_hex(0, 0, 0) == "000000"

    def test_white(self):
        assert _pixel_to_hex(255, 255, 255) == "FFFFFF"


class TestColLetter:
    def test_single_letter(self):
        assert _col_letter(1) == "A"
        assert _col_letter(26) == "Z"

    def test_double_letter(self):
        assert _col_letter(27) == "AA"
        assert _col_letter(52) == "AZ"


class TestIm2xlsx:
    def test_creates_xlsx(self, tiny_image: Path):
        result = im2xlsx(tiny_image, resize=False)
        assert result.suffix == ".xlsx"
        assert result.exists()

    def test_returns_path(self, tiny_image: Path):
        result = im2xlsx(tiny_image, resize=False)
        assert isinstance(result, Path)

    def test_cell_colors_match_pixels(self, tiny_image: Path):
        result = im2xlsx(tiny_image, resize=False)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None

        # Top-left pixel should be red
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb == "00FF0000"

        # Top-middle pixel should be green
        cell = ws.cell(row=1, column=2)
        assert cell.fill.start_color.rgb == "0000FF00"

    def test_dimensions_match_image(self, tiny_image: Path):
        result = im2xlsx(tiny_image, resize=False)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 2
        assert ws.max_column == 3

    def test_resize_shrinks_large_image(self, large_image: Path):
        result = im2xlsx(large_image, resize=True)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None
        assert ws.max_row is not None
        assert ws.max_column is not None
        assert ws.max_row <= MAX_HEIGHT
        assert ws.max_column <= MAX_WIDTH

    def test_resize_with_keep_aspect(self, large_image: Path):
        result = im2xlsx(large_image, resize=True, keep_aspect=True)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None
        assert ws.max_row is not None
        assert ws.max_column is not None
        assert ws.max_row <= MAX_HEIGHT
        assert ws.max_column <= MAX_WIDTH

    def test_no_resize_keeps_original_size(self, large_image: Path):
        result = im2xlsx(large_image, resize=False)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 600
        assert ws.max_column == 500

    def test_file_not_found(self, tmp_path: Path):
        with pytest.raises(FileNotFoundError):
            im2xlsx(tmp_path / "nonexistent.png")

    def test_zoom_level_is_low(self, tiny_image: Path):
        result = im2xlsx(tiny_image, resize=False)
        wb = load_workbook(result)
        ws = wb.active
        assert ws is not None
        assert ws.sheet_view.zoomScale == 10
